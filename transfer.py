#This for Huawei Packing List Only!!!
import openpyxl

def copy_cells(source_file, source_sheet, target_file, target_sheet):
    count_total = 0
    true_no = 0
    # Load the source workbook and sheet
    source_workbook = openpyxl.load_workbook(source_file)
    source_sheet = source_workbook[source_sheet]

    # Load the target workbook and sheet
    target_workbook = openpyxl.load_workbook(target_file)
    target_sheet = target_workbook[target_sheet]

    # Read values from source columns F and J, starting from row 7
    target_sheet.delete_rows(1, target_sheet.max_row)  # Clear existing data in target sheet
    for row in source_sheet.iter_rows(min_row=7, min_col=6, values_only=True):  # F corresponds to column 6
        f_value = row[0]  # Column F
        if f_value is None or f_value.split(" ") == "":
            break
        repeat_sn = 0
        check_no = False
        unavailable = "NA"
        print(f_value)
        if row[4] is None or row[4].split(" ")== "":
            j_values = unavailable
        elif row[4] is not None and row[4]!= '':
            j_values = row[4].split(" ")  # Column J values separated by space

        # Paste values into target columns A and B
        if j_values == unavailable or j_values[0] == '':
            repeat_sn += 1
            count_total += 1
            check_no = True
            j_values = unavailable
            target_sheet.cell(row=target_sheet.max_row + 1, column=1, value=f_value)  # Column A
            target_sheet.cell(row=target_sheet.max_row, column=2, value=j_values)  # Column B
        else:
            for j_value in j_values:
                repeat_sn += 1
                count_total += 1
                target_sheet.cell(row=target_sheet.max_row + 1, column=1, value=f_value)  # Column A    
                target_sheet.cell(row=target_sheet.max_row, column=2, value=j_value)  # Column B
        if check_no == True:
            true_no += 1
            print("CONFIRM THIS QUANTITY")
        print("Quantity: ", repeat_sn,"\n")
    # Save the target workbook
    target_workbook.save(target_file)

    print("Cells copied successfully!")
    print("Total Quantity: ", count_total,"\n")
    print("Items with no SN: (NOT QUANTITY)", true_no,"\n")
    print("Total Quantity (With SN): ", count_total-true_no,"\n")

# ONLY EDIT HERE
# Example usage
source_file_name = "i60688.xlsx"
source_file = "packing_list\\"+source_file_name
source_sheet = "PACKING LIST"
target_file = "result_SN\sn.xlsx"
target_sheet = "Sheet1"


copy_cells(source_file, source_sheet, target_file, target_sheet)