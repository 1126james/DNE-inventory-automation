from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests

template_path = 'DNE_Template.xlsx'
sn_file_path = 'sn.xlsx'
workbook_template = load_workbook(filename=template_path)
workbook_sn = load_workbook(filename=sn_file_path)

# Assuming the data is in cell A1
data_template = workbook_template['A1'].value
data_sn = workbook_sn['A1'].value

url = 'https://example.com/form'
response = requests.get(url)
html = response.text
soup = BeautifulSoup(html, 'html.parser')
# Find the input field (replace 'input_field_id' with the actual ID or other attribute of the input field)
input_field = soup.find('input', {'id': 'input_field_id'})

# Clear the input field (optional, remove if not needed)
input_field['value'] = ''

# Input the data from the Excel file
input_field['value'] = data
# Find the form element (replace 'form_id' with the actual ID or other attribute of the form)
form = soup.find('form', {'id': 'form_id'})

# Submit the form (replace 'submit_button_id' with the actual ID or other attribute of the submit button)
submit_button = form.find('button', {'id': 'submit_button_id'})
response = requests.post(url, data=form)
workbook.close()